from django.shortcuts import render

# Create your views here.

from django.http import HttpResponse
from rest_framework.views import APIView
import openpyxl
from .models import shweta


#Django REST framework and is used to define the behavior of HTTP methods (like GET, PO

class Fileapi(APIView):
    def post(self, request):
        file=request.data.get('file')
        workbook=openpyxl.load_workbook(file)
        # sheet = workbook.sheetnames[0]
        # sheet_data = workbook[sheet]
        # print(sheet_data)
        sheet=workbook.active

        

        for row in sheet.iter_rows(min_row=2, values_only=True):
                name, age = row
                shweta.objects.create(name=name, age=age)



        return HttpResponse("okay")
